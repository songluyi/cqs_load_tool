# -*- coding: utf-8 -*-
# 2016/7/15 13:36
"""
-------------------------------------------------------------------------------
Function:   database of cqs_index
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import cx_Oracle
from openpyxl import load_workbook

def get_indexid():
    conn = cx_Oracle.connect("apps/apps@192.168.15.94:1539/NRCRP2")
    cur =conn.cursor()
    r= cur.execute("select t.index_id from CUX.CUX_CQS_INDEX_T t")
    index_id=[]
    for row in cur:
        index_id.extend(row)
    print('数据库中最大的Index_ID为')
    print(max(index_id))
    return max(index_id)

def get_batch_id():
    conn = cx_Oracle.connect("apps/apps@192.168.15.94:1539/NRCRP2")
    cur =conn.cursor()
    r= cur.execute("select t.batch_id from CUX.CUX_CQS_INDEX_T t")
    result=cur.fetchone()
    new_result=list(result)
    return new_result[0]
def insert_db(row):
    conn = cx_Oracle.connect("apps/apps@192.168.15.94:1539/NRCRP2")
    cur =conn.cursor()
    r= cur.execute(" INSERT INTO CUX.CUX_CQS_INDEX_T values (:INDEX_ID,:INDEX_ORDER,:CATAGORY,:SERVICES,:DESIGN_TEMP_SOURCE,:DESIGN_TEMP_MIN,:DESIGN_TEMP_MAX,:DESIGN_TEMP_SPEC,:DESIGN_PRES_SOURCE,:DESIGN_PRES_MIN,:DESIGN_PRES_MAX,:DESIGN_PRES_SPEC,:PIPING_MATL_CLASS,:BASIC_MATERIAL,:RATING,:FLANGE_FACING,:CA,:NOTE,:BATCH_ID,:CREATED_BY,TO_DATE(:CREATION_DATE,'yyyy/mm/dd'),:LAST_UPDATED_BY,TO_DATE(:LAST_UPDATE_DATE,'yyyy/mm/dd'),:LAST_UPDATE_LOGIN,:ATTRIBUTE_CATEGORY,:ATTRIBUTE1,:ATTRIBUTE2,:ATTRIBUTE3,:ATTRIBUTE4,:ATTRIBUTE5,:ATTRIBUTE6,:ATTRIBUTE7,:ATTRIBUTE8,:ATTRIBUTE9,:ATTRIBUTE10,:ATTRIBUTE11,:ATTRIBUTE12,:ATTRIBUTE13,:ATTRIBUTE14,:ATTRIBUTE15)", row)
    conn.commit()
def compliment():
    header_name=['INDEX_ID', 'INDEX_ORDER', 'CATAGORY', 'SERVICES', 'DESIGN_TEMP_SOURCE', 'DESIGN_TEMP_MIN',
             'DESIGN_TEMP_MAX', 'DESIGN_TEMP_SPEC', 'DESIGN_PRES_SOURCE', 'DESIGN_PRES_MIN',
             'DESIGN_PRES_MAX', 'DESIGN_PRES_SPEC', 'PIPING_MATL_CLASS', 'BASIC_MATERIAL', 'RATING',
             'FLANGE_FACING', 'CA', 'NOTE', 'BATCH_ID', 'CREATED_BY', 'CREATION_DATE', 'LAST_UPDATED_BY',
             'LAST_UPDATE_DATE', 'LAST_UPDATE_LOGIN', 'ATTRIBUTE_CATEGORY', 'ATTRIBUTE1', 'ATTRIBUTE2',
             'ATTRIBUTE3', 'ATTRIBUTE4', 'ATTRIBUTE5', 'ATTRIBUTE6', 'ATTRIBUTE7', 'ATTRIBUTE8', 'ATTRIBUTE9',
             'ATTRIBUTE10', 'ATTRIBUTE11', 'ATTRIBUTE12', 'ATTRIBUTE13', 'ATTRIBUTE14', 'ATTRIBUTE15']
    name='new索引表.xlsx'
    wb_get_excel = load_workbook(filename=name)
    sheets = wb_get_excel.get_sheet_names()
    ws_get_excel = wb_get_excel.get_sheet_by_name(sheets[0])
    line=2
    while ws_get_excel.cell(row=line, column=1).value is not None:
       excel_data=[]
       for column in range(1,len(header_name)+1):#循环取列值
           excel_data.append(ws_get_excel.cell(row=line, column=column).value)
           if  isinstance(ws_get_excel.cell(row=line, column=column).value, int):#判定值如果为int类型改为float更符合oracle
               excel_data[column-1]=float(excel_data[column-1])
       make_dict=dict(zip(header_name,excel_data))
       print(make_dict)
       print(len(make_dict))
       row=tuple(excel_data)
       new_row=[]
       new_row.append(row)
       line=line+1
       insert_db(make_dict)
       print('已经插入数据，行数为',line)
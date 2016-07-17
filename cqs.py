# -*- coding: utf-8 -*-
# 2016/7/14 19:26
"""
-------------------------------------------------------------------------------
Function:
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import os
# import cx_Oracle
from openpyxl import Workbook
from openpyxl import load_workbook
load_name=r'管道材料等级索引表.xlsx'
import time
today_time=time.strftime("%Y-%m-%d", time.localtime())
today_time=today_time.replace('-','/')
header_name=['INDEX_ID', 'INDEX_ORDER', 'CATAGORY', 'SERVICES', 'DESIGN_TEMP_SOURCE', 'DESIGN_TEMP_MIN',
             'DESIGN_TEMP_MAX', 'DESIGN_TEMP_SPEC', 'DESIGN_PRES_SOURCE', 'DESIGN_PRES_MIN',
             'DESIGN_PRES_MAX', 'DESIGN_PRES_SPEC', 'PIPING_MATL_CLASS', 'BASIC_MATERIAL', 'RATING',
             'FLANGE_FACING', 'CA', 'NOTE', 'BATCH_ID', 'CREATED_BY', 'CREATION_DATE', 'LAST_UPDATED_BY',
             'LAST_UPDATE_DATE', 'LAST_UPDATE_LOGIN', 'ATTRIBUTE_CATEGORY', 'ATTRIBUTE1', 'ATTRIBUTE2',
             'ATTRIBUTE3', 'ATTRIBUTE4', 'ATTRIBUTE5', 'ATTRIBUTE6', 'ATTRIBUTE7', 'ATTRIBUTE8', 'ATTRIBUTE9',
             'ATTRIBUTE10', 'ATTRIBUTE11', 'ATTRIBUTE12', 'ATTRIBUTE13', 'ATTRIBUTE14', 'ATTRIBUTE15', 'ROWID']
class cqs(object):
    def make_exceldata(self):
        wb_write=Workbook()
        ws_write = wb_write.get_active_sheet()
        ws_write.title = 'PMC Index (Sorted by Services)'
        #写入基本的excel表头
        for i in range(1,41):
            ws_write.cell(row=1, column=i).value = header_name[i]
        wb_load = load_workbook(filename=load_name)
        sheets = wb_load.get_sheet_names()
        ws_load = wb_load.get_sheet_by_name(sheets[0])
        #获取catalog在excel的行数,在样例中输出为[7,9,13,30]
        catolog_line=[]
        for j in range(1,1000):
            if str(ws_load.cell(row=j+6, column=1).value)>3:
               line_num=j+6
               catolog_line.append(line_num)
        line=0
        for count in range(0,len(catolog_line)-1):
            for row in range(catolog_line[count],catolog_line[count+1]):
                line+=1
                ws_write.cell(row=line, column=2).value=ws_load.cell(row=row+1, column=1) #写入INDEX_ORDER
                ws_write.cell(row=line, column=3).value=ws_load.cell(row=row, column=1) #写入catalog
                ws_write.cell(row=line, column=4).value=ws_load.cell(row=row+1, column=2) #写入SERVICES
                ws_write.cell(row=line, column=5).value=ws_load.cell(row=row+1, column=3) #写入DESIGN_TEMP_SOURCE 跳过了需要选择判断的内容，该内容用函数调用生成
                ws_write.cell(row=line, column=9).value=ws_load.cell(row=row+1, column=4) #写入DESIGN_PRES_SOURCE
                ws_write.cell(row=line, column=13).value=ws_load.cell(row=row+1, column=5) #写入PIPING_MATL_CLASS
                ws_write.cell(row=line, column=14).value=ws_load.cell(row=row+1, column=6) #写入BASIC_MATERIAL
                ws_write.cell(row=line, column=15).value=ws_load.cell(row=row+1, column=7) #写入RATING
                ws_write.cell(row=line, column=16).value=ws_load.cell(row=row+1, column=8) #写入FLANGE_FACING
                ws_write.cell(row=line, column=17).value=ws_load.cell(row=row+1, column=9) #写入CA
                ws_write.cell(row=line, column=19).value=0.00 #写入BATCH_ID note值不导入 因此留空
                ws_write.cell(row=line, column=20).value=-1 #写入CREATED_BY
                ws_write.cell(row=line, column=21).value=today_time #写入CREATION_DATE
                ws_write.cell(row=line, column=22).value=-1 #写入LAST_UPDATED_BY
                ws_write.cell(row=line, column=23).value=today_time #写入CREATION_DATE
                ws_write.cell(row=line, column=24).value=-1 #写入LAST_UPDATE_LOGIN
                ws_write.cell(row=line, column=25).value=today_time #写入CREATION_DATE
                ws_write.cell(row=line, column=41).value='AADbvzAApAAAAFdAAA' #写入ROWID
        wb_write.save('1.xlsx')



cqs=cqs()
cqs.make_exceldata()

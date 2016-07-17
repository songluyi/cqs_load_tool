# -*- coding: utf-8 -*-
# 2016/7/15 15:06
"""
-------------------------------------------------------------------------------
Function:   用于将管道等级表从excel直接插入到plsql里面
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
# -*- coding: utf-8 -*-
# 2016/7/14 19:26
import os
# import cx_Oracle
from openpyxl import Workbook
from openpyxl import load_workbook
import time
from database import get_indexid
today_time=time.strftime("%Y-%m-%d", time.localtime())
today_time=today_time.replace('-','/')
header_name=['INDEX_ID', 'INDEX_ORDER', 'CATAGORY', 'SERVICES', 'DESIGN_TEMP_SOURCE', 'DESIGN_TEMP_MIN',
             'DESIGN_TEMP_MAX', 'DESIGN_TEMP_SPEC', 'DESIGN_PRES_SOURCE', 'DESIGN_PRES_MIN',
             'DESIGN_PRES_MAX', 'DESIGN_PRES_SPEC', 'PIPING_MATL_CLASS', 'BASIC_MATERIAL', 'RATING',
             'FLANGE_FACING', 'CA', 'NOTE', 'BATCH_ID', 'CREATED_BY', 'CREATION_DATE', 'LAST_UPDATED_BY',
             'LAST_UPDATE_DATE', 'LAST_UPDATE_LOGIN', 'ATTRIBUTE_CATEGORY', 'ATTRIBUTE1', 'ATTRIBUTE2',
             'ATTRIBUTE3', 'ATTRIBUTE4', 'ATTRIBUTE5', 'ATTRIBUTE6', 'ATTRIBUTE7', 'ATTRIBUTE8', 'ATTRIBUTE9',
             'ATTRIBUTE10', 'ATTRIBUTE11', 'ATTRIBUTE12', 'ATTRIBUTE13', 'ATTRIBUTE14', 'ATTRIBUTE15', 'ROWID']
class cqs(object):
    def make_exceldata(self,name,index_id):
        wb_write=Workbook()
        ws_write = wb_write.get_active_sheet()
        ws_write.title = 'PMC Index'
        #写入基本的excel表头
        for i in range(1,42):
            ws_write.cell(row=1, column=i).value = header_name[i-1]
        wb_load = load_workbook(filename=load_name)
        sheets = wb_load.get_sheet_names()
        ws_load = wb_load.get_sheet_by_name(sheets[0])
        #获取catalog在excel的行数,在样例中输出为[7,9,13,30]
        catolog_line=[]
        for j in range(1,1000):
            if '类' in str(ws_load.cell(row=j+6, column=1).value):
               line_num=j+6
               catolog_line.append(line_num)
        line=1
        for count in range(0,len(catolog_line)-1):
            for row in range(catolog_line[count],catolog_line[count+1]-1):#这里for循环我到时候会用yield来重写，for看起来很蠢
                line+=1
                ws_write.cell(row=line, column=1).value=int(index_id)+1#ID先设定为空
                index_id+=1                                            #自增1
                ws_write.cell(row=line, column=2).value=ws_load.cell(row=row+1, column=1).value #写入INDEX_ORDER
                ws_write.cell(row=line, column=3).value=ws_load.cell(row=catolog_line[count], column=1).value #写入catalog
                ws_write.cell(row=line, column=4).value=ws_load.cell(row=row+1, column=2).value #写入SERVICES
                ws_write.cell(row=line, column=5).value=ws_load.cell(row=row+1, column=3).value #写入DESIGN_TEMP_SOURCE 跳过了需要选择判断的内容，该内容用函数调用生成
                ws_write.cell(row=line, column=9).value=ws_load.cell(row=row+1, column=4).value #写入DESIGN_PRES_SOURCE
                ws_write.cell(row=line, column=13).value=ws_load.cell(row=row+1, column=5).value #写入PIPING_MATL_CLASS
                ws_write.cell(row=line, column=14).value=ws_load.cell(row=row+1, column=6).value #写入BASIC_MATERIAL
                ws_write.cell(row=line, column=15).value=ws_load.cell(row=row+1, column=7).value #写入RATING
                ws_write.cell(row=line, column=16).value=ws_load.cell(row=row+1, column=8).value #写入FLANGE_FACING
                ws_write.cell(row=line, column=17).value=ws_load.cell(row=row+1, column=9).value #写入CA
                ws_write.cell(row=line, column=19).value=0.00 #写入BATCH_ID note值不导入 因此留空
                ws_write.cell(row=line, column=20).value=-1 #写入CREATED_BY
                ws_write.cell(row=line, column=21).value=today_time #写入CREATION_DATE
                ws_write.cell(row=line, column=22).value=-1 #写入LAST_UPDATED_BY
                ws_write.cell(row=line, column=23).value=today_time #写入CREATION_DATE
                ws_write.cell(row=line, column=24).value=-1 #写入LAST_UPDATE_LOGIN
                ws_write.cell(row=line, column=25).value=today_time #写入CREATION_DATE
                ws_write.cell(row=line, column=41).value='AADbvzAApAAAAFdAAA' #写入ROWID

                #开始利用正则来处理TEMP和PRES
                temp_data=ws_load.cell(row=row+1, column=3).value
                if '≤' in temp_data:
                    new_data=temp_data.replace('≤','')
                    ws_write.cell(row=line, column=7).value=new_data #读入到最大的max temp里面
                elif '~' in temp_data:
                    new_data=temp_data.split('~')
                    ws_write.cell(row=line, column=6).value=new_data[0]#小的读入min temp里面
                    ws_write.cell(row=line, column=7).value=new_data[1]#小的读入max temp里面
                else:
                    ws_write.cell(row=line, column=8).value=temp_data#其他无法解析情况扔进spec temp里面

                pres_data=ws_load.cell(row=row+1, column=4).value
                if ',' in pres_data:
                    new_data=pres_data.split(',')
                    ws_write.cell(row=line, column=10).value=new_data[0].replace('>','')#小的读入min temp里面
                    ws_write.cell(row=line, column=11).value=new_data[1].replace('≤','')#小的读入max temp里面
                elif '≤' in pres_data:
                    new_data=pres_data.replace('≤','')
                    ws_write.cell(row=line, column=11).value=new_data#小的读入max temp里面
                else:
                    ws_write.cell(row=line, column=12).value=pres_data#其他无法解析情况扔进spec temp里面

        name='new'+name
        wb_write.save(name)
        print('已经完成提交')

    def get_path(self):
        import os
        path=os.getcwd()
        FileList = []
        rootdir = path

        for root, subFolders, files in os.walk(rootdir):
            #排除特定的子目录
            if 'done' in subFolders:
                subFolders.remove('done')
            #查找特定扩展名的文件，只要包含'.xlsx'字符串的文件，都会被包含
            for f in files:
                if f.find('.xlsx') != -1:
                    FileList.append(os.path.join(root, f))

        for item in FileList:
            print('检测到您目录下有如下excel文件 请确保他们是要提交的管道等级索引表')
            print(item)
        return FileList


if __name__ == '__main__':
    cqs=cqs()
    name_list=cqs.get_path()
    for name in name_list:
        load_name=os.path.basename(name)
        index_id=get_indexid()

        cqs.make_exceldata(load_name,index_id)

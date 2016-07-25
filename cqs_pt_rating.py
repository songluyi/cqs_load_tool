# -*- coding: utf-8 -*-
# 2016/7/19 14:40
"""
-------------------------------------------------------------------------------
Function:   处理等级表中的pt_rating
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from cqs_pt_rating_database import *
import multiprocessing
from concurrent import futures#采用3.x新出来的多进程
import time
today_time=time.strftime("%Y-%m-%d", time.localtime())
today_time=today_time.replace('-','/')
header_name=['PT_ID', 'BATCH_ID', 'PT_ORDER_NUMBER', 'PIPING_MATL_CLASS', 'TEMPERATURE', 'PRESSURE',
             'CREATED_BY', 'CREATION_DATE', 'LAST_UPDATED_BY', 'LAST_UPDATE_DATE', 'LAST_UPDATE_LOGIN',
             'ATTRIBUTE_CATEGORY', 'ATTRIBUTE1', 'ATTRIBUTE2', 'ATTRIBUTE3', 'ATTRIBUTE4', 'ATTRIBUTE5',
             'ATTRIBUTE6', 'ATTRIBUTE7', 'ATTRIBUTE8', 'ATTRIBUTE9', 'ATTRIBUTE10', 'ATTRIBUTE11',
             'ATTRIBUTE12','ATTRIBUTE13', 'ATTRIBUTE14', 'ATTRIBUTE15']

class cqs_pt_rating(object):
   def get_path(self):
        import os
        path=os.getcwd()
        FileList = []
        rootdir = path

        for root, subFolders, files in os.walk(rootdir):
            #排除特定的子目录
            if 'done' in subFolders:
                subFolders.remove('done')
            #查找特定扩展名的文件，只要包含'索引表'但不包含"new"字符串的文件，都会被包含
            for f in files:
                if f.find('等级表') != -1 and f.find('new')==-1:
                    FileList.append(os.path.join(root, f))

        for item in FileList:
            print('检测到您目录下有如下excel文件 请确保他们是要提交的管道等级表')
            print(item)
        return FileList

   def make_exceldata(self,name_list,bug_pi_id,pi_id,batch_id,pt_order_number):
        wb_write=Workbook()
        ws_write = wb_write.get_active_sheet()
        ws_write.title = 'pt_rating'
        line=1#从第二行开始
        batch_id+=1
        for i in range(1,28):
                ws_write.cell(row=1, column=i).value = header_name[i-1]

        for name in name_list:
            load_name=os.path.basename(name)
            if bug_pi_id>pi_id:#因为读取不同的excel，此时数据库中还没有同步插入 必须要用一个bug_pi_id来解决下一个文件的pi_id的取值问题
                pi_id=bug_pi_id
            wb_load = load_workbook(filename=load_name)
            sheets = wb_load.get_sheet_names()

            del sheets[-1]#用这个模块尾部会多一个页签
            for page_name in sheets[::2]:#循环页签
                ws_load = wb_load.get_sheet_by_name(page_name)

                for i in range(3,21):#判定表格中第九行三到而是一行是否为数字是就执行插入数据

                    if ws_load.cell(row=9, column=i).value is not None:
                        pi_id+=1
                        pt_order_number+=1
                        line+=1
                        bug_pi_id+=1
                        ws_write.cell(row=line, column=1).value=pi_id
                        ws_write.cell(row=line, column=2).value=batch_id
                        ws_write.cell(row=line, column=3).value=pt_order_number
                        ws_write.cell(row=line, column=4).value=ws_load.cell(row=5, column=5).value#写入metal class
                        ws_write.cell(row=line, column=5).value=ws_load.cell(row=9, column=i).value#写入温度
                        ws_write.cell(row=line, column=6).value=ws_load.cell(row=10, column=i).value
                        ws_write.cell(row=line, column=7).value=0#写入created by
                        ws_write.cell(row=line, column=8).number_format='yyyy-mm-dd'
                        ws_write.cell(row=line, column=8).value=today_time
                        ws_write.cell(row=line, column=9).value=0#写入last_update_by
                        ws_write.cell(row=line, column=10).number_format='yyyy-mm-dd'
                        ws_write.cell(row=line, column=10).value=today_time
                        ws_write.cell(row=line, column=11).value=0#写入last_update_login
        name='new'+'压力温度'+'.xlsx'
        wb_write.save(name)
        print('已经完成压力温度表的excel生成')
        return bug_pi_id
if __name__ == '__main__':
    # ex = futures.ThreadPoolExecutor(max_workers=4)
    pool_size=multiprocessing.cpu_count()*2
    pool=multiprocessing.Pool(processes=pool_size)
    cqs=cqs_pt_rating()
    name_list=cqs.get_path()
    bug_pi_id=0
    pi_id=get_ptid()
    space_tab=0
    batch_id=get_batch_id()
    pt_order_number=get_order_number()
    cqs.make_exceldata(name_list,bug_pi_id,pi_id,batch_id,pt_order_number)
    excel_name='new压力温度.xlsx'
    data_list=compliment(header_name,excel_name)
    result=pool.map(insert_db,data_list)
    print('最后已经完成提交cqs的压力与温度数据提交~谢谢使用')

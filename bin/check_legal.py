# -*- coding: utf-8 -*-
# 2016/8/3 18:14
"""
-------------------------------------------------------------------------------
Function:
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
from about_ftp import get_path
from openpyxl import load_workbook
import cqs_index_database
import cqs_items_database
import cqs_note_database
import cqs_pipe_thickness_database
import cqs_pt_rating_database
import cqs_branch_connect_database
import time
def check_batch_id(check_flag=0):
    index_batch_id=cqs_index_database.get_batch_id()
    item_batch_id=cqs_items_database.get_batch_id()
    note_batch_id=cqs_note_database.get_batch_id()
    pipe_batch_id=cqs_pipe_thickness_database.get_batch_id()
    pt_rate_batch_id=cqs_pt_rating_database.get_batch_id()
    connect_batch_id=cqs_branch_connect_database.get_batch_id()
    if check_flag==0:
        print('管道材料等级索引表历史最大批次号',index_batch_id)
        print('管道材料等级表-元件表历史最大批次号',item_batch_id)
        print('管道材料等级索引表/等级表-备注内容历史最大批次号',note_batch_id)
        print('管道材料等级表-外径壁厚表历史最大批次号',pipe_batch_id)
        print('管道材料等级表-压力温度表历史最大批次号',pt_rate_batch_id)
        print('管道材料等级表-支管连接表最大批次号',connect_batch_id)
    else:
        print('管道材料等级索引表历史最大批次号',index_batch_id-1)
        print('管道材料等级表-元件表历史最大批次号',item_batch_id-1)
        print('管道材料等级索引表/等级表-备注内容历史最大批次号',note_batch_id-1)
        print('管道材料等级表-外径壁厚表历史最大批次号',pipe_batch_id-1)
        print('管道材料等级表-压力温度表历史最大批次号',pt_rate_batch_id-1)
        print('管道材料等级表-支管连接表最大批次号',connect_batch_id-1)
    if index_batch_id==item_batch_id==note_batch_id==pipe_batch_id==pt_rate_batch_id==connect_batch_id:
        flag=True
        print('目前所有批次号正常')
        return [flag]
    else:
        print('error:批次号可能存在异常，原因可能是导入异常数据导致，请进入PLSQL界面操作，保证批次号相同')
        flag=False
        time.sleep(5)
        id_list=[index_batch_id,item_batch_id,note_batch_id,pipe_batch_id,pt_rate_batch_id,connect_batch_id]
        return [flag,id_list]


if __name__ == '__main__':
    namelist=get_path()
    print(namelist)
    index_error_time=0
    rate_error_time=0
    check_double=0
    if len(namelist)==0:
        print('error:excel目录下为空，请将需要导入的excel表格放置其中!')
        rate_error_time+=1
        index_error_time+=1
    for name in namelist:
        if '索引表' in name:
            check_double+=1
            wb_load = load_workbook(filename=name)
            sheets = wb_load.get_sheet_names()
            ws_load = wb_load.get_sheet_by_name(sheets[0])

            if '适用介质' in ws_load.cell(row=5, column=2).value and '设计温度' in ws_load.cell(row=5, column=3).value\
                    and '设计压力' in ws_load.cell(row=5, column=4).value and '管道材料' in ws_load.cell(row=5, column=5).value\
                    and '基本材料' in ws_load.cell(row=5, column=6).value and '压力' in ws_load.cell(row=5, column=7).value\
                    and '法兰' in ws_load.cell(row=5,column=8).value and '腐蚀裕量' in ws_load.cell(row=5,column=9).value\
                    and '备注' in ws_load.cell(row=5,column=10).value:
                print('该页签下第五行管道材料索引表符合导入格式要求')
            else:
                index_error_time+=1
                print('\nerror:管道材料索引表第五行 不 符合导入格式要求！')
                time.sleep(2)
            if 'Services' in ws_load.cell(row=6, column=2).value and 'Temp' in ws_load.cell(row=6, column=3).value\
                    and 'Pres' in ws_load.cell(row=6, column=4).value and 'Piping' in ws_load.cell(row=6, column=5).value\
                    and 'Material' in ws_load.cell(row=6, column=6).value and 'Rating' in ws_load.cell(row=6, column=7).value\
                    and 'Flange' in ws_load.cell(row=6,column=8).value and 'C. A.' in ws_load.cell(row=6,column=9).value\
                    and 'Note' in ws_load.cell(row=6,column=10).value:
                print('该页签下第六行管道材料索引表符合导入格式要求')
            else:
                index_error_time+=1
                print('\nerror:管道材料索引表第六行 不 符合导入格式要求！')
            if index_error_time>0:
                print('\nerror:管道材料索引表不符合要求，请关闭本程序修改索引表')

        error_page_name=[]
        if '等级表' in name:
            check_double+=1
            wb_load = load_workbook(filename=name)
            sheets = wb_load.get_sheet_names()
            del sheets[-1]#用openpyxl这个模块获取excel标签尾部会多一个页签，具体原因不明

            for page_name in sheets[::2]:#循环页签
                ws_load = wb_load.get_sheet_by_name(page_name)
                if '温度' in ws_load.cell(row=9, column=1).value and '压力' in ws_load.cell(row=10, column=1).value:
                    print('管道材料等级表-压力温度表格式检测正常')
                else:
                    err_msg='管道材料等级表-压力温度表格式出错,在页签 '+str(page_name)
                    rate_error_time+=1
                    error_page_name.append(err_msg)
                if '元件名称' in ws_load.cell(row=11,column=1).value and '公称直径' in ws_load.cell(row=11,column=4).value\
                        and '端面' in ws_load.cell(row=11,column=8).value and '壁厚' in ws_load.cell(row=11,column=10).value\
                        and '材料' in ws_load.cell(row=11,column=12).value and '标准' in ws_load.cell(row=11,column=16).value\
                        and '备注' in ws_load.cell(row=11,column=20).value and '最小' in ws_load.cell(row=12,column=4).value\
                        and '最大' in ws_load.cell(row=12,column=6).value:
                    print('该页签项下管道材料等级表-元件表格式检查正常')
                else:
                    err_msg='管道材料等级表-元件表格式出错,在页签 '+str(page_name)
                    rate_error_time+=1
                    error_page_name.append(err_msg)
            for page_name in sheets[1::2]:#循环页签
                ws_load = wb_load.get_sheet_by_name(page_name)
                if 'DN' in ws_load.cell(row=5, column=1).value and '外径' in ws_load.cell(row=6, column=1).value\
                        and 'DN' in ws_load.cell(row=8, column=1).value and '外径' in ws_load.cell(row=9, column=1).value\
                        and 'DN' in ws_load.cell(row=11, column=1).value and '外径' in ws_load.cell(row=12, column=1).value:
                    print('该页签项下管道材料等级表-外径壁厚表格式检查正常')
                else:
                    err_msg='管道材料等级表-外径壁厚表格式出错,在页签 '+str(page_name)
                    rate_error_time+=1
                    error_page_name.append(err_msg)
                if '主管' in ws_load.cell(row=45,column=5).value:
                    print('该页签项下管道材料等级表-支管连接表格式检查正常')
                else:
                    err_msg='管道材料等级表-支管连接表格式出错,在页签 '+str(page_name)
                    rate_error_time+=1
                    error_page_name.append(err_msg)
            if rate_error_time>0:
                for single_msg in error_page_name:
                    print(single_msg)
                print('\nerror:等级表导入数据错误')
            else:
                print('等级表符合导入要求')
                time.sleep(3)
    if check_double % 2>0:
        print('\nerror:管道材料索引表必须和其等级表匹配放置!')
        rate_error_time+=1
        time.sleep(3)#停留三秒方便用户查看
    while rate_error_time>0 or index_error_time>0:
        try:
            check_flag=input('\n原始excel数据检测不符合导入条件，不允许导入!请立刻关闭本程序，并修改excel：')
        except ValueError:
            print("Oops!  Please close this function thx")
    check_batch_id()
# -*- coding: utf-8 -*-
# 2016/7/28 16:41
"""
-------------------------------------------------------------------------------
Function:
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import time
from cqs_note_database import *
from cqs_pt_rating import cqs_pt_rating#这里主要为了引入get_path函数 我将生成新excel数据读入成list封装成一个函数
from cqs_pt_rating_database import compliment
from openpyxl.styles import numbers
import time
header_name=['NOTE_ID', 'SERVICE_SOURCE', 'NOTE_KEY', 'NOTE', 'BATCH_ID', 'CREATED_BY', 'CREATION_DATE',
             'LAST_UPDATED_BY', 'LAST_UPDATE_DATE', 'LAST_UPDATE_LOGIN', 'ATTRIBUTE_CATEGORY',
             'ATTRIBUTE1', 'ATTRIBUTE2', 'ATTRIBUTE3', 'ATTRIBUTE4', 'ATTRIBUTE5', 'ATTRIBUTE6',
             'ATTRIBUTE7', 'ATTRIBUTE8', 'ATTRIBUTE9', 'ATTRIBUTE10', 'ATTRIBUTE11', 'ATTRIBUTE12',
             'ATTRIBUTE13', 'ATTRIBUTE14', 'ATTRIBUTE15']
today_time=time.strftime("%Y-%m-%d", time.localtime())
today_time=today_time.replace('-','/')
class cqs_note(object):
    def make_exceldata(self,name_list,bug_note_id,note_id,batch_id):
        wb_write=Workbook()
        ws_write = wb_write.get_active_sheet()
        ws_write.title = 'note'
        line=1#从第二行开始
        note_id=note_id+1
        batch_id=batch_id+1
        for i in range(1,27):#写入头文件
                ws_write.cell(row=1, column=i).value = header_name[i-1]
        for name in name_list:
            load_name=os.path.basename(name)
            if '等级表'in load_name:
                if bug_note_id>note_id:#因为读取不同的excel，此时数据库中还没有同步插入 必须要用一个bug_note_id来解决下一个文件的pi_id的取值问题
                    note_id=bug_note_id
                wb_load = load_workbook(filename=name)
                sheets = wb_load.get_sheet_names()
                del sheets[-1]#用openpyxl这个模块获取excel标签尾部会多一个页签，具体原因不明

                for page_name in sheets[::2]:#循环页签
                    ws_load = wb_load.get_sheet_by_name(page_name)
                    note_line=500
                    for i in range(1,500):
                        if '备注：' in str(ws_load.cell(row=i, column=1).value):
                            note_line=i
                    while ws_load.cell(row=note_line+1, column=1).value is not None and note_line<500:
                            line+=1
                            bug_note_id+=1
                            ws_write.cell(row=line, column=1).value=note_id#写入item_id
                            note_id+=1
                            if ws_load.cell(row=note_line+2, column=1).value is None:
                                row_str=str(ws_load.cell(row=note_line+1, column=1).value)+str(ws_load.cell(row=note_line+2, column=2).value)
                            else:
                                row_str=ws_load.cell(row=note_line+1, column=1).value
                            ws_write.cell(row=line, column=2).value=ws_load.cell(row=5, column=5).value#写入service source
                            if ord(row_str[5])>46:
                                NOTE_KEY='注'+str(row_str[4])+str(row_str[5])
                            else:
                                NOTE_KEY='注'+str(row_str[4])
                            ws_write.cell(row=line, column=3).value=NOTE_KEY#写入note_key

                            ws_write.cell(row=line, column=4).value=row_str#写入note信息
                            ws_write.cell(row=line, column=5).value=batch_id#写入batch_id
                            ws_write.cell(row=line, column=6).value=0#写入created by
                            ws_write.cell(row=line, column=7).value=today_time#写入creat_time
                            ws_write.cell(row=line, column=8).value=0#update_by
                            ws_write.cell(row=line, column=9).value=today_time#写入Last_date
                            ws_write.cell(row=line, column=10).value=0#写入update_login
                            note_line+=1
                            print('写入了一个到excel')
            elif '索引表' in load_name:
                if bug_note_id>note_id:#因为读取不同的excel，此时数据库中还没有同步插入 必须要用一个bug_note_id来解决下一个文件的pi_id的取值问题
                    note_id=bug_note_id
                wb_load = load_workbook(filename=name)
                sheets = wb_load.get_sheet_names()
                del sheets[-1]#用openpyxl这个模块获取excel标签尾部会多一个页签，具体原因不明
                #这里刚好不用重复,写一个就好
                ws_load = wb_load.get_sheet_by_name(sheets[0])
                count=499
                for i in range(1,500):
                    if '备注：' in str(ws_load.cell(row=i, column=1).value):
                        count=i
                for note_line in range(count,500):
                    if ws_load.cell(row=note_line+1, column=1).value is not None:
                            line+=1
                            bug_note_id+=1
                            ws_write.cell(row=line, column=1).value=note_id#写入item_id
                            note_id+=1
                            ws_write.cell(row=line, column=2).value='INDEXER'
                            if ws_load.cell(row=note_line+2, column=1).value is None and ws_load.cell(row=note_line+2, column=2).value is not None:
                                row_str=str(ws_load.cell(row=note_line+1, column=1).value)+str(ws_load.cell(row=note_line+1, column=2).value)+str(ws_load.cell(row=note_line+2, column=2).value)
                            else:
                                row_str=str(ws_load.cell(row=note_line+1, column=1).value)+str(ws_load.cell(row=note_line+1, column=2).value)
                            # row_str=str(row_str)+str(ws_load.cell(row=note_line+1, column=2).value)
                            if ord(row_str[1])>60:
                                NOTE_KEY='注'+str(row_str[0])
                            else:
                                NOTE_KEY='注'+str(row_str[0])+str(row_str[1])
                            ws_write.cell(row=line, column=3).value=NOTE_KEY#写入note_key
                            ws_write.cell(row=line, column=4).value=row_str#写入note信息
                            ws_write.cell(row=line, column=5).value=batch_id#写入batch_id
                            ws_write.cell(row=line, column=6).value=0#写入created by
                            ws_write.cell(row=line, column=7).value=today_time#写入creat_time
                            ws_write.cell(row=line, column=8).value=0#update_by
                            ws_write.cell(row=line, column=9).value=today_time#写入Last_date
                            ws_write.cell(row=line, column=10).value=0#写入update_login
                            note_line+=1
                            print('写入了一个到excel')

        name='new'+'注释表'+'.xlsx'
        wb_write.save(name)
        print('已经生成excel，请注意查看根目录')

    def get_path(self):
        import os
        current_path=os.path.abspath(os.path.join(os.path.dirname('cqs_index.py'),os.path.pardir))
        new_path=current_path+'\\'+'excel\\'
        FileList = []
        rootdir = new_path
        for root, subFolders, files in os.walk(rootdir):
            #排除特定的子目录
            if 'done' in subFolders:
                subFolders.remove('done')
            #查找特定扩展名的文件，只要包含'索引表'但不包含"new"字符串的文件，都会被包含
            for f in files:
                if f.find('表') != -1 and f.find('new')==-1:
                    FileList.append(os.path.join(root, f))
        for item in FileList:
            print('检测到您目录下有如下excel文件 请确保他们是要提交的管道等级表')
            print(item)
        return FileList
if __name__ == '__main__':
    cqs=cqs_note()
    name_list=cqs.get_path()
    note_id=get_noteid()
    if note_id is None:
        note_id=0
    batch_id=get_batch_id()
    if batch_id is None:
        batch_id=0
    bug_note_id=0
    cqs.make_exceldata(name_list,bug_note_id,note_id,batch_id)
    excel_name='new注释表.xlsx'
    data_list=compliment(header_name,excel_name)
    start_time=time.time()
    insert_db(data_list)
    end_time=time.time()
    print('耗时为：',end_time-start_time,'插入总数为：',len(data_list))
    print('已经完成对管道注释表表的插入，谢谢使用')



# -*- coding: utf-8 -*-
# 2016/7/20 13:38
"""
-------------------------------------------------------------------------------
Function:
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import cx_Oracle,random
from openpyxl import load_workbook
from cqs_pt_rating_database import get_ini
db_connect=get_ini()[0]
def get_pipeid():
    conn = cx_Oracle.connect(db_connect)
    cur =conn.cursor()
    r= cur.execute("select max(pipe_id) from cux.cux_cqs_pipe_thickness_his_t")
    result=cur.fetchone()
    new_result=list(result)
    return new_result[0]
def get_batch_id():
    conn = cx_Oracle.connect(db_connect)
    cur =conn.cursor()
    r= cur.execute("select max(batch_id) from cux.cux_cqs_pipe_thickness_his_t")
    result=cur.fetchone()
    new_result=list(result)
    return new_result[0]
def get_order_number():
    conn = cx_Oracle.connect(db_connect)
    cur =conn.cursor()
    r= cur.execute("select max(pipe_order_number) from cux.cux_cqs_pipe_thickness_his_t")
    result=cur.fetchone()
    new_result=list(result)
    return new_result[0]
def insert_db(row):
    conn = cx_Oracle.connect(db_connect)
    cur =conn.cursor()
    r=cur.execute("truncate table cux.cux_cqs_pipe_thickness_t")
    r= cur.executemany(" INSERT INTO cux.cux_cqs_pipe_thickness_t values (:PIPE_ID,:BATCH_ID,:PIPE_ORDER_NUMBER,:PIPING_MATL_CLASS,:PIPE_DN,:PIPE_OUTER,:PIPE_THICKNESS,:CREATED_BY,to_date(:CREATION_DATE,'yyyy/mm/dd'),:LAST_UPDATED_BY,to_date(:LAST_UPDATE_DATE,'yyyy/mm/dd'),:LAST_UPDATE_LOGIN,:ATTRIBUTE_CATEGORY,:ATTRIBUTE1,:ATTRIBUTE2,:ATTRIBUTE3,:ATTRIBUTE4,:ATTRIBUTE5,:ATTRIBUTE6,:ATTRIBUTE7,:ATTRIBUTE8,:ATTRIBUTE9,:ATTRIBUTE10,:ATTRIBUTE11,:ATTRIBUTE12,:ATTRIBUTE13,:ATTRIBUTE14,:ATTRIBUTE15)", row)
    r=cur.execute("insert into  cux.cux_cqs_pipe_thickness_his_t select * from cux.cux_cqs_pipe_thickness_t")
    conn.commit()
    print('数据已经导入成功')

# def compliment(header_name,name):
#     header_name=header_name
#     wb_get_excel = load_workbook(filename=name)
#     sheets = wb_get_excel.get_sheet_names()
#     ws_get_excel = wb_get_excel.get_sheet_by_name(sheets[0])
#     line=2
#     full_shit_list=[]
#     while ws_get_excel.cell(row=line, column=1).value is not None:
#        excel_data=[]
#        for column in range(1,len(header_name)+1):#循环取列值
#            excel_data.append(ws_get_excel.cell(row=line, column=column).value)
#            if  isinstance(ws_get_excel.cell(row=line, column=column).value, int):#判定值如果为int类型改为float更符合oracle
#                excel_data[column-1]=float(excel_data[column-1])
#        make_dict=dict(zip(header_name,excel_data))
#        line=line+1
#        full_shit_list.append(make_dict)
#     return full_shit_list



# -*- coding: utf-8 -*-
# 2016/7/19 19:27
"""
-------------------------------------------------------------------------------
Function:   将索引表里的元件数据导入数据库
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import time
from cqs_items_database import *
from cqs_pt_rating import cqs_pt_rating#这里主要为了引入get_path函数
from cqs_pt_rating_database import compliment
from openpyxl.styles import numbers
import time
header_name=['ITEM_ID', 'BATCH_ID', 'ITEM_ORDER_NUMBER', 'PIPING_MATL_CLASS', 'ITEM_CATEGORY',
             'ITEM_NAME', 'MIN_DN', 'MAX_DN', 'END_FACING', 'THK_RATING', 'MATERIAL', 'STANDARD_MODEL',
             'NOTE', 'CREATED_BY', 'CREATION_DATE', 'LAST_UPDATED_BY', 'LAST_UPDATE_DATE',
             'LAST_UPDATE_LOGIN', 'ATTRIBUTE_CATEGORY', 'ATTRIBUTE1', 'ATTRIBUTE2', 'ATTRIBUTE3',
             'ATTRIBUTE4', 'ATTRIBUTE5', 'ATTRIBUTE6', 'ATTRIBUTE7', 'ATTRIBUTE8', 'ATTRIBUTE9',
             'ATTRIBUTE10', 'ATTRIBUTE11', 'ATTRIBUTE12', 'ATTRIBUTE13', 'ATTRIBUTE14', 'ATTRIBUTE15']
today_time=time.strftime("%Y-%m-%d", time.localtime())
today_time=today_time.replace('-','/')
class cqs_items(object):
    def make_exceldata(self,name_list,bug_item_id,item_id,batch_id,item_order_number):
        wb_write=Workbook()
        ws_write = wb_write.get_active_sheet()
        ws_write.title = 'item'
        line=1#从第二行开始
        item_id=item_id+1
        batch_id=batch_id+1
        item_order_number=item_order_number+1
        for i in range(1,35):#写入头文件
                ws_write.cell(row=1, column=i).value = header_name[i-1]
        for name in name_list:
            # load_name=os.path.basename(name)
            if bug_item_id>item_id:#因为读取不同的excel，此时数据库中还没有同步插入 必须要用一个bug_pi_id来解决下一个文件的pi_id的取值问题
                item_id=bug_item_id
            wb_load = load_workbook(filename=name)
            sheets = wb_load.get_sheet_names()
            del sheets[-1]#用openpyxl这个模块获取excel标签尾部会多一个页签，具体原因不明
            for page_name in sheets[::2]:#循环页签
                ws_load = wb_load.get_sheet_by_name(page_name)
                catolog_line=[]
                for i in range(13,500):
                    if '备注' in str(ws_load.cell(row=i, column=1).value):
                        od=i-1#向上提一行
                for i in range(13,od):
                    if ws_load.cell(row=i, column=1).value is not None and ws_load.cell(row=i, column=4).value is None:
                        catolog_line.append(i)
                catolog_line.append(od)
                print(catolog_line)
                first_one=catolog_line[0]
                last_one=catolog_line[-1]
                nothing_id=last_one-first_one-len(catolog_line)+1
                for count in range(0,len(catolog_line)-1):#这里是类别循环 [13, 16, 20, 35, 38, 40, 42, 56] 样本一共六组
                    for row in range(catolog_line[count]+1,catolog_line[count+1]):
                        if ws_load.cell(row=row, column=4).value is not None:
                            line+=1
                            bug_item_id+=1
                            ws_write.cell(row=line, column=1).value=item_id#写入item_id
                            item_id+=1
                            ws_write.cell(row=line, column=2).value=batch_id#写入batch_id
                            ws_write.cell(row=line, column=3).value=item_order_number#写入item order number
                            item_order_number+=1
                            ws_write.cell(row=line, column=4).value=ws_load.cell(row=5, column=5).value#写入pip_matel_class
                            ws_write.cell(row=line, column=5).value=str(ws_load.cell(row=catolog_line[count], column=1).value)#写入item_catalog
                            ws_write.cell(row=line, column=6).value=str(ws_load.cell(row=row, column=1).value)#写入item_name
                            ws_write.cell(row=line, column=7).value=ws_load.cell(row=row, column=4).value#写入DN min
                            ws_write.cell(row=line, column=8).value=ws_load.cell(row=row, column=6).value#写入DN max
                            ws_write.cell(row=line, column=9).value=ws_load.cell(row=row, column=8).value#写入END_FACING
                            ws_write.cell(row=line, column=10).value=ws_load.cell(row=row, column=10).value#写入THK_RATING
                            # print(ws_load.cell(row=row, column=12).value)
                            # print(type(ws_load.cell(row=row, column=12).value))
                            if ws_load.cell(row=row, column=12).value is None:
                               ws_write.cell(row=line, column=11).value=''
                            else:
                                ws_write.cell(row=line, column=11).value=str(ws_load.cell(row=row, column=12).value)#写入MATERIAL
                            ws_write.cell(row=line, column=12).value=ws_load.cell(row=row, column=16).value#写入STANDARD_MODEL
                            ws_write.cell(row=line, column=13).value=ws_load.cell(row=row, column=20).value#写入STANDARD_MODEL
                            ws_write.cell(row=line, column=14).value=0#写入created by
                            ws_write.cell(row=line, column=15).value=today_time
                            ws_write.cell(row=line, column=16).value=0#写入update by
                            ws_write.cell(row=line, column=17).value=today_time
                            ws_write.cell(row=line, column=18).value=0#写入last__update__login
        name='new'+'管道材料等级表-元件表'+'.xlsx'
        wb_write.save(name)
        print('已经生成excel，请注意查看根目录')
if __name__ == '__main__':
    cqs=cqs_items()
    name_list=cqs_pt_rating().get_path()
    item_id=get_itemid()
    if item_id is None:
        item_id=0
    batch_id=get_batch_id()
    if batch_id is None:
        batch_id=0
    item_order_number=get_order_number()
    if item_order_number is None:
        item_order_number=0
    bug_item_id=0
    cqs.make_exceldata(name_list,bug_item_id,item_id,batch_id,item_order_number)
    excel_name='new管道材料等级表-元件表.xlsx'
    data_list=compliment(header_name,excel_name)
    start_time=time.time()
    insert_db(data_list)
    end_time=time.time()
    print('耗时为：',end_time-start_time,'插入总数为：',len(data_list))
    print('已经完成对管道元件表的插入，谢谢使用')



# -*- coding: utf-8 -*-
# 2016/7/20 15:36
"""
-------------------------------------------------------------------------------
Function:   支管连接表插入数据库主程序
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import os,sys
from openpyxl import Workbook
from openpyxl import load_workbook
from check_legal import check_batch_id
from cqs_branch_connect_database import *
from cqs_pt_rating import cqs_pt_rating#这里主要为了引入get_path函数
from cqs_pt_rating_database import compliment
import time
import argparse
parser = argparse.ArgumentParser(description='choose new_load_data like zero or load_data like one.')
parser.add_argument('-t', action="store", dest="choice_value", type=int)
args = parser.parse_args()
today_time=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
today_time=today_time.replace('-','/')
header_name=['CONNECTION_ID', 'BATCH_ID', 'CONN_ORDER_NUMBER', 'PIPING_MATL_CLASS', 'BRANCH_DN',
             'HEADER_DN', 'CONNECTION_TYPE', 'CREATED_BY', 'CREATION_DATE', 'LAST_UPDATED_BY',
             'LAST_UPDATE_DATE', 'LAST_UPDATE_LOGIN', 'ATTRIBUTE_CATEGORY', 'ATTRIBUTE1', 'ATTRIBUTE2',
             'ATTRIBUTE3', 'ATTRIBUTE4', 'ATTRIBUTE5', 'ATTRIBUTE6', 'ATTRIBUTE7', 'ATTRIBUTE8',
             'ATTRIBUTE9', 'ATTRIBUTE10', 'ATTRIBUTE11', 'ATTRIBUTE12', 'ATTRIBUTE13', 'ATTRIBUTE14',
             'ATTRIBUTE15']
class cqs_branch_connect(object):

    def make_exceldata(self,name_list,bug_connection_id,connection_id,batch_id,conn_order_number,domain_name):
        wb_write=Workbook()
        ws_write = wb_write.get_active_sheet()
        ws_write.title = 'pt_rating'
        line=1#从第二行开始
        connection_id=connection_id+1
        batch_id=batch_id+1
        conn_order_number=conn_order_number+1
        for i in range(1,29):
                ws_write.cell(row=1, column=i).value = header_name[i-1]
        for name in name_list:
            # load_name=os.path.basename(name)
            if bug_connection_id>connection_id:#因为读取不同的excel，此时数据库中还没有同步插入 必须要用一个bug_pipe_id来解决下一个文件的pipe_id的取值问题
                connection_id=bug_connection_id
            wb_load = load_workbook(filename=name)
            sheets = wb_load.get_sheet_names()
            del sheets[-1]#用openpyxl这个模块获取excel标签尾部会多一个页签，具体原因不明
            for page_name in sheets[1::2]:#循环样例应该为16.2 35.2 等
                ws_load = wb_load.get_sheet_by_name(page_name)
                check_row_lines=[]
                for row in range(14,45):
                    if isinstance(ws_load.cell(row=row, column=3).value,int):
                        check_row_lines.append(row)
                check_row=min(check_row_lines)
                max_column=5+max(check_row_lines)-check_row
                for column in range(5,max_column+1):
                    count=column-4
                    for row in range(check_row+count-1,44):
                        connection_id+=1
                        conn_order_number+=1
                        line+=1
                        bug_connection_id+=1
                        ws_write.cell(row=line, column=1).value=connection_id
                        ws_write.cell(row=line, column=2).value=batch_id
                        ws_write.cell(row=line, column=3).value=conn_order_number
                        ws_write.cell(row=line, column=4).value=ws_load.cell(row=1, column=26).value#写入metal class
                        ws_write.cell(row=line, column=5).value=ws_load.cell(row=44, column=column).value   #写入主管
                        # print(row,column)
                        ws_write.cell(row=line, column=6).value=ws_load.cell(row=row, column=3).value   #写入支管
                        ws_write.cell(row=line, column=7).value=ws_load.cell(row=row, column=column).value  #连接方式
                        ws_write.cell(row=line, column=8).value=domain_name#写入created by
                        ws_write.cell(row=line, column=9).number_format='yyyy-mm-dd'
                        ws_write.cell(row=line, column=9).value=today_time
                        ws_write.cell(row=line, column=10).value=0#写入last_update_by
                        ws_write.cell(row=line, column=11).number_format='yyyy-mm-dd'
                        ws_write.cell(row=line, column=11).value=today_time
                        ws_write.cell(row=line, column=12).value=0#写入last_update_login
        name='new'+'管道材料等级表-支管连接表'+'.xlsx'
        wb_write.save(name)
        print('已经完成管道材料等级表-支管连接表的excel生成')

# def check_continue():
#     while True:
#         try:
#             print('\n下面将会对导入的数据进行检查，在未返回主界面之前请不要关闭本窗口')
#             check_flag=int(input('\n请输入一个数字，0表示正常类型，1表示续传类型：'))
#             if -1<check_flag<2:
#                 break
#             else:
#                 print('尽可以允许输入0和1')
#         except ValueError:
#             print("Oops!  That was no valid number.  Try again...")
#     return check_flag

if __name__ == '__main__':
    cqs=cqs_branch_connect()
    name_list=cqs_pt_rating().get_path()
    connection_id=get_connectionid()
    if connection_id is None:
        connection_id=0
    batch_id=get_batch_id()
    if batch_id is None:
        batch_id=0
    conn_order_number=get_order_number()
    if conn_order_number is None:
        conn_order_number=0
    bug_connection_id=0
    domain_name=return_domain_username()
    cqs.make_exceldata(name_list,bug_connection_id,connection_id,batch_id,conn_order_number,domain_name)
    excel_name='new管道材料等级表-支管连接表.xlsx'
    data_list=compliment(header_name,excel_name)
    start_time=time.time()
    insert_db(data_list)
    end_time=time.time()
    print('耗时为：',end_time-start_time,'插入总数为：',len(data_list))
    print('已经完成对管道材料等级表-支管连接表的插入，谢谢使用')
    print('如果未出现完成插入的提示则说明该表导入失败')
    time.sleep(10)
    check_flag=args.choice_value
    print(check_flag)
    flag=check_batch_id(check_flag=check_flag)[0]
    if flag is False:
        print('本次导入失败，部分等级表无法被写入数据库，将执行删除操作')
        min_batch_id=min(check_batch_id()[1])
        sql_commands=make_sql_str(min_batch_id)
        print(sql_commands)
        s=list(map(del_rest_data,sql_commands))
        print('删除成功')
    else:
        insert_batch_db(today_time)







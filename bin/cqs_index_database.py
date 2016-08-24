# -*- coding: utf-8 -*-
# 2016/7/15 13:36
"""
-------------------------------------------------------------------------------
Function:   database of cqs_index
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import cx_Oracle,os
from openpyxl import load_workbook
from cqs_pt_rating_database import get_ini
db_connect=get_ini()[0]
# 设置编码，否则：
# 1. Oracle 查询出来的中文是乱码
# 2. 插入数据时有中文，会导致
# UnicodeEncodeError: 'ascii' codec can't encode characters in position 1-7: ordinal not in range(128)
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
def get_indexid():
    conn = cx_Oracle.connect(db_connect)
    cur =conn.cursor()
    r= cur.execute("select max(index_id) from CUX.CUX_CQS_INDEX_HIS_T")
    result=cur.fetchone()
    new_result=list(result)
    print('数据库中最大的INDEX_ID为')
    print(new_result[0])
    return new_result[0]

def get_batch_id():
    conn = cx_Oracle.connect(db_connect)
    cur =conn.cursor()
    r= cur.execute("select max(batch_id) from CUX.CUX_CQS_INDEX_HIS_T")
    result=cur.fetchone()
    new_result=list(result)
    # print('数据库中最大的batch_ID为')
    # print(new_result[0])
    return new_result[0]

def insert_db(row):
    conn = cx_Oracle.connect(db_connect)
    cur =conn.cursor()
    r=cur.execute("truncate table cux.cux_cqs_index_t")
    r=cur.executemany(" INSERT INTO CUX.CUX_CQS_INDEX_T values (:INDEX_ID,:INDEX_ORDER,:CATAGORY,:SERVICES,:DESIGN_TEMP_SOURCE,:DESIGN_TEMP_MIN,:DESIGN_TEMP_MAX,:DESIGN_TEMP_SPEC,:DESIGN_PRES_SOURCE,:DESIGN_PRES_MIN,:DESIGN_PRES_MAX,:DESIGN_PRES_SPEC,:PIPING_MATL_CLASS,:BASIC_MATERIAL,:RATING,:FLANGE_FACING,:CA,:NOTE,:BATCH_ID,:CREATED_BY,TO_DATE(:CREATION_DATE,'yyyy/mm/dd'),:LAST_UPDATED_BY,TO_DATE(:LAST_UPDATE_DATE,'yyyy/mm/dd'),:LAST_UPDATE_LOGIN,:ATTRIBUTE_CATEGORY,:ATTRIBUTE1,:ATTRIBUTE2,:ATTRIBUTE3,:ATTRIBUTE4,:ATTRIBUTE5,:ATTRIBUTE6,:ATTRIBUTE7,:ATTRIBUTE8,:ATTRIBUTE9,:ATTRIBUTE10,:ATTRIBUTE11,:ATTRIBUTE12,:ATTRIBUTE13,:ATTRIBUTE14,:ATTRIBUTE15)", row)
    r=cur.execute("insert into  cux.cux_cqs_index_his_t select * from cux.cux_cqs_index_t")
    conn.commit()

def compliment():
    header_name=['INDEX_ID', 'INDEX_ORDER', 'CATAGORY', 'SERVICES', 'DESIGN_TEMP_SOURCE', 'DESIGN_TEMP_MIN',
             'DESIGN_TEMP_MAX', 'DESIGN_TEMP_SPEC', 'DESIGN_PRES_SOURCE', 'DESIGN_PRES_MIN',
             'DESIGN_PRES_MAX', 'DESIGN_PRES_SPEC', 'PIPING_MATL_CLASS', 'BASIC_MATERIAL', 'RATING',
             'FLANGE_FACING', 'CA', 'NOTE', 'BATCH_ID', 'CREATED_BY', 'CREATION_DATE', 'LAST_UPDATED_BY',
             'LAST_UPDATE_DATE', 'LAST_UPDATE_LOGIN', 'ATTRIBUTE_CATEGORY', 'ATTRIBUTE1', 'ATTRIBUTE2',
             'ATTRIBUTE3', 'ATTRIBUTE4', 'ATTRIBUTE5', 'ATTRIBUTE6', 'ATTRIBUTE7', 'ATTRIBUTE8', 'ATTRIBUTE9',
             'ATTRIBUTE10', 'ATTRIBUTE11', 'ATTRIBUTE12', 'ATTRIBUTE13', 'ATTRIBUTE14', 'ATTRIBUTE15']
    name='new管道材料等级索引表.xlsx'
    # excel_path=os.getcwd()+'\\'+'bin\\'+name
    wb_get_excel = load_workbook(filename=name)
    sheets = wb_get_excel.get_sheet_names()
    ws_get_excel = wb_get_excel.get_sheet_by_name(sheets[0])
    line=2
    full_shit_list=[]
    while ws_get_excel.cell(row=line, column=1).value is not None:
       excel_data=[]
       for column in range(1,len(header_name)+1):#循环取列值
           excel_data.append(ws_get_excel.cell(row=line, column=column).value)
           if  isinstance(ws_get_excel.cell(row=line, column=column).value, int):#判定值如果为int类型改为float更符合oracle
               excel_data[column-1]=float(excel_data[column-1])
       make_dict=dict(zip(header_name,excel_data))
       line=line+1
       full_shit_list.append(make_dict)
    return full_shit_list
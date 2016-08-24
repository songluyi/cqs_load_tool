# -*- coding: utf-8 -*-
# 2016/7/15 15:06
"""
-------------------------------------------------------------------------------
Function:   用于将管道等级表从excel直接插入到数据库里面
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""
import os,random
from openpyxl import Workbook
from openpyxl import load_workbook
import time
from cqs_index_database import get_indexid,compliment,get_batch_id,insert_db
from openpyxl.styles import numbers
style_index_id=numbers.NumberFormatDescriptor(style=numbers.FORMAT_NUMBER_00)
today_time=time.strftime("%Y-%m-%d", time.localtime())
today_time=today_time.replace('-','/')
header_name=['INDEX_ID', 'INDEX_ORDER', 'CATAGORY', 'SERVICES', 'DESIGN_TEMP_SOURCE', 'DESIGN_TEMP_MIN',
             'DESIGN_TEMP_MAX', 'DESIGN_TEMP_SPEC', 'DESIGN_PRES_SOURCE', 'DESIGN_PRES_MIN',
             'DESIGN_PRES_MAX', 'DESIGN_PRES_SPEC', 'PIPING_MATL_CLASS', 'BASIC_MATERIAL', 'RATING',
             'FLANGE_FACING', 'CA', 'NOTE', 'BATCH_ID', 'CREATED_BY', 'CREATION_DATE', 'LAST_UPDATED_BY',
             'LAST_UPDATE_DATE', 'LAST_UPDATE_LOGIN', 'ATTRIBUTE_CATEGORY', 'ATTRIBUTE1', 'ATTRIBUTE2',
             'ATTRIBUTE3', 'ATTRIBUTE4', 'ATTRIBUTE5', 'ATTRIBUTE6', 'ATTRIBUTE7', 'ATTRIBUTE8', 'ATTRIBUTE9',
             'ATTRIBUTE10', 'ATTRIBUTE11', 'ATTRIBUTE12', 'ATTRIBUTE13', 'ATTRIBUTE14', 'ATTRIBUTE15', 'ROWID']
class cqs_index(object):
    def make_exceldata(self,bug_index_id,name_list,space_tab,index_id,nothing_id,batch_id):
        wb_write=Workbook()
        ws_write = wb_write.get_active_sheet()
        ws_write.title = 'PMC Index'
        for name in name_list:
            space_tab=space_tab+nothing_id
            # load_name=os.path.basename(name)
            if bug_index_id>index_id:#因为读取不同的excel，此时数据库中还没有同步插入 必须要用一个bug_pi_id来解决下一个文件的pi_id的取值问题
                index_id=bug_index_id
            #写入基本的excel表头
            for i in range(1,42):
                ws_write.cell(row=1, column=i).value = header_name[i-1]
            wb_load = load_workbook(filename=name)
            sheets = wb_load.get_sheet_names()
            ws_load = wb_load.get_sheet_by_name(sheets[0])
            #获取catalog在excel的行数,在样例中输出为[7,9,13,30,38]
            catolog_line=[]
            bug_index_id=index_id
            for i in range(1,1000):
                if '备注' in str(ws_load.cell(row=i, column=1).value):
                    od=i-1#向上提一行
            for j in range(1,od):
                if '类' in str(ws_load.cell(row=j+6, column=1).value):
                   line_num=j+6
                   catolog_line.append(line_num)
            catolog_line.append(od)
            first_one=catolog_line[0]
            last_one=catolog_line[-1]
            nothing_id=last_one-first_one-len(catolog_line)+1
            line=1
            for count in range(0,len(catolog_line)-1):
                for row in range(catolog_line[count],catolog_line[count+1]-1):#根据确定的行和列遍历
                    line+=1
                    bug_index_id+=1
                    edit_index_id=int(index_id)+1
                    ws_write.cell(row=line+space_tab, column=1).number_format='0.00'#设置单元格格式
                    ws_write.cell(row=line+space_tab, column=1).value=edit_index_id#ID先设定为空
                    index_id+=1                                            #自增1
                    ws_write.cell(row=line+space_tab, column=2).number_format='0.00'
                    ws_write.cell(row=line+space_tab, column=2).value=ws_load.cell(row=row+1, column=1).value #写入INDEX_ORDER
                    ws_write.cell(row=line+space_tab, column=3).value=ws_load.cell(row=catolog_line[count], column=1).value #写入catalog
                    ws_write.cell(row=line+space_tab, column=4).value=ws_load.cell(row=row+1, column=2).value #写入SERVICES
                    ws_write.cell(row=line+space_tab, column=5).value=ws_load.cell(row=row+1, column=3).value #写入DESIGN_TEMP_SOURCE
                    ws_write.cell(row=line+space_tab, column=9).value=ws_load.cell(row=row+1, column=4).value #写入DESIGN_PRES_SOURCE
                    ws_write.cell(row=line+space_tab, column=13).value=ws_load.cell(row=row+1, column=5).value #写入PIPING_MATL_CLASS
                    ws_write.cell(row=line+space_tab, column=14).value=ws_load.cell(row=row+1, column=6).value #写入BASIC_MATERIAL
                    ws_write.cell(row=line+space_tab, column=15).value=ws_load.cell(row=row+1, column=7).value #写入RATING
                    ws_write.cell(row=line+space_tab, column=16).value=ws_load.cell(row=row+1, column=8).value #写入FLANGE_FACING
                    ws_write.cell(row=line+space_tab, column=17).number_format='0.00'
                    ws_write.cell(row=line+space_tab, column=17).value=float(ws_load.cell(row=row+1, column=9).value) #写入CA
                    ws_write.cell(row=line+space_tab, column=18).value=ws_load.cell(row=row+1, column=10).value#写入note注释
                    ws_write.cell(row=line+space_tab, column=19).value=batch_id #写入BATCH_ID
                    ws_write.cell(row=line+space_tab, column=20).value=-1 #写入CREATED_BY
                    ws_write.cell(row=line+space_tab, column=21).number_format='yyyy-mm-dd'
                    ws_write.cell(row=line+space_tab, column=21).value=today_time #写入CREATION_DATE
                    ws_write.cell(row=line+space_tab, column=22).number_format='0.00'
                    ws_write.cell(row=line+space_tab, column=22).value=-1 #写入LAST_UPDATED_BY
                    ws_write.cell(row=line+space_tab, column=23).number_format='yyyy-mm-dd'
                    ws_write.cell(row=line+space_tab, column=23).value=today_time #写入CREATION_DATE
                    ws_write.cell(row=line+space_tab, column=24).number_format='0.00'
                    ws_write.cell(row=line+space_tab, column=24).value=-1 #写入LAST_UPDATE_LOGIN
                    # ws_write.cell(row=line, column=25).value=today_time #写入CREATION_DATE
                    ws_write.cell(row=line+space_tab, column=41).value='AADbvzAApAAAAFdAAA' #写入ROWID

                    #开始利用正则来处理TEMP和PRES
                    temp_data=ws_load.cell(row=row+1, column=3).value
                    if temp_data is not None:
                        # print(temp_data)
                        ws_write.cell(row=line+space_tab, column=7).number_format='0.00'
                        ws_write.cell(row=line+space_tab, column=6).number_format='0.00'
                        ws_write.cell(row=line+space_tab, column=10).number_format='0.00'
                        ws_write.cell(row=line+space_tab, column=12).number_format='0.00'
                        ws_write.cell(row=line+space_tab, column=11).number_format='0.00'
                        if '≤' in temp_data:
                            new_data=temp_data.replace('≤','')
                            ws_write.cell(row=line+space_tab, column=7).value=float(new_data) #读入到最大的max temp里面
                        elif '~' in temp_data:
                            new_data=temp_data.split('~')
                            ws_write.cell(row=line+space_tab, column=6).value=float(new_data[0])#小的读入min temp里面
                            ws_write.cell(row=line+space_tab, column=7).value=float(new_data[1])#小的读入max temp里面
                        else:
                            ws_write.cell(row=line+space_tab, column=8).value=float(temp_data)#其他无法解析情况扔进spec temp里面

                    pres_data=ws_load.cell(row=row+1, column=4).value
                    if pres_data is not None:
                        if pres_data=="常压":
                            ws_write.cell(row=line+space_tab, column=10).value=0
                            ws_write.cell(row=line+space_tab, column=11).value=1
                            ws_write.cell(row=line+space_tab, column=12).value=pres_data
                        if ',' in pres_data:
                            new_data=pres_data.split(',')
                            ws_write.cell(row=line+space_tab, column=10).value=float(new_data[0].replace('＞',''))#小的读入min press里面
                            ws_write.cell(row=line+space_tab, column=11).value=float(new_data[1].replace('≤',''))#小的读入max press里面
                        elif '≤' in pres_data:
                            new_data=pres_data.replace('≤','')
                            ws_write.cell(row=line+space_tab, column=11).value=float(new_data)#小的读入max press里面
                        else:
                            ws_write.cell(row=line+space_tab, column=12).value=pres_data#其他无法解析情况扔进spec press里面
        name='new'+'管道材料等级索引表'+'.xlsx'
        wb_write.save(name)
        print('已经生成管道材料等级索引表excel，请注意查看根目录')
        return index_id
#获取当前目录下的xlsx文件
    def get_path(self):
        import os
        current_path=os.path.abspath(os.path.join(os.path.dirname('cqs_index.py'),os.path.pardir))
        new_path=current_path+'\\'+'excel\\'
        FileList = []
        rootdir = new_path

        for root, subFolders, files in os.walk(rootdir):
            #排除特定的子目录
            if 'done' in subFolders:
                subFolders.remove('done')
            #查找特定扩展名的文件，只要包含'索引表'但不包含"new"字符串的文件，都会被包含
            for f in files:
                if f.find('索引表') != -1 and f.find('new')==-1:
                    FileList.append(os.path.join(root, f))

        for item in FileList:
            print('检测到您目录下有如下excel文件 请确保他们是要提交的管道等级索引表')
            print(item)
        return FileList



if __name__ == '__main__':
    start_time=time.time()
    cqs=cqs_index()
    name_list=cqs.get_path()
    bug_index_id=0
    space_tab=0
    index_id=get_indexid()
    if index_id is None:
        index_id=0
    nothing=0
    batch_id=get_batch_id()
    if batch_id is None:
        batch_id=0
    batch_id=batch_id+1
    bug_index_id=cqs.make_exceldata(bug_index_id,name_list,space_tab,index_id,nothing,batch_id)
    data_list=compliment()
    insert_db(data_list)
    end_time=time.time()
    print('最后已经完成提交管道材料等级索引表~谢谢使用')
    print('耗时：')
    print(end_time-start_time)
    print('如果未出现完成插入的提示则说明该表导入失败')
    time.sleep(5)